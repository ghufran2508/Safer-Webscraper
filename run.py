import gspread
from oauth2client.service_account import ServiceAccountCredentials
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
# import pprint

is_google_drive = False
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

def extract_info(mcn):
    url = 'https://safer.fmcsa.dot.gov/query.asp'

    form_data =  {'searchtype': 'ANY', 'query_type': 'queryCarrierSnapshot', 'query_param': 'MC_MX', 'query_string': mcn}

    session = requests.Session()
    response = session.post(url, headers=headers, data=form_data)

    print(response)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all <tr> elements
        trs = soup.find_all('tr')

        # Find the <tr> element whose <th> has class 'querylabelbkg'
        target_trs = [tr for tr in trs if tr.find('th', class_='querylabelbkg')]

        if len(target_trs) == 0:
            print("No Data found")
            return None
        else:
            # chech Entity Type
            if target_trs[0].find('td', class_='queryfield').text.strip() == "CARRIER":
                if target_trs[1].find('td', class_='queryfield').text.strip() == "ACTIVE":
                    usdot = target_trs[2].find('td', class_='queryfield').text.strip()
                    authority_status = target_trs[4].find('td', class_='queryfield').text.split('  ')[0]

                    if authority_status in ["NOT FOUND", "NOT AUTHORIZED", "OUT OF SERVICE"]:
                        return None

                    company_name = target_trs[6].find('td', class_='queryfield').text.strip()
                    address = target_trs[8].find('td', class_='queryfield').text.strip()
                    phone = target_trs[9].find('td', class_='queryfield').text.strip()
                    email = 'NOT FOUND'

                    url2 = f'https://ai.fmcsa.dot.gov/SMS/Carrier/{usdot}/CarrierRegistration.aspx'
                    response2 = session.get(url2, headers=headers)

                    if response2.status_code == 200:
                        soup2 = BeautifulSoup(response2.content, 'html.parser')
                        email_div = soup2.find('div', class_='modalBody')

                        if email_div:
                            spans = email_div.findAll('span', class_='dat')
                            if len(spans) > 7:
                                email = spans[6].text.strip()

                    return company_name, address, phone, email, authority_status, usdot                        
    else:
        print("Error: Failed to retrieve data")

    return None

def main():
    
    creds_file = 'client_key.json'
    file_name = ""

    if os.path.isfile(creds_file):
        file_name = "fmcsa_data"
        print("Authenticating Google Drive....")
        #Authorize the API
        scope = [
            'https://www.googleapis.com/auth/drive',
            'https://www.googleapis.com/auth/drive.file'
            ]
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_file,scope)
        client = gspread.authorize(credentials=creds)

        print("Authentication complete.")
        print('Opening file...')

        global is_google_drive
        is_google_drive=True

        #Fetch the sheet
        sheet = client.open(file_name).sheet1
        # python_sheet = sheet.get_all_records()
        # pp = pprint.PrettyPrinter()
        # pp.pprint(python_sheet)
    else:
        file_name = "fmcsa_data.xlsx"
        if os.path.isfile(file_name):
            wb = openpyxl.load_workbook(filename=file_name)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "FMCSA Data"

            headers = ["MC Number", "US Dot", "Company Name", "Phone Number", "Email", "Physical Address", "Authority Status"]
            sheet.append(headers)

    start_range = int(input("Enter the starting MC number: "))
    end_range = int(input("Enter the ending MC number: "))

    for mc_number in range(start_range, end_range + 1):
        data = extract_info(mc_number)
        if data:
            company_name, address, phone, email, authority_status, usdot = data
            if is_google_drive:
                sheet.append_row([mc_number, usdot, company_name, phone, email, address, authority_status])
            else:
                sheet.append([mc_number, usdot, company_name, phone, email, address, authority_status])
        
        print(mc_number, " data extraction completed....")
        if not is_google_drive:
            wb.save(filename=file_name)

    if not is_google_drive:
        wb.close()
        print(f"Data extracted and saved to {file_name}")
    else:
        print("Data updated to google drive")


if __name__ == "__main__":
    main()



# #Fetch row
# row = sheet.row_values(5)
# print('\nFetched Row')
# pp.pprint(row)
# #Fetch column
# col = sheet.col_values(2)
# print('\nFetched Column')
# pp.pprint(col)
# #Fetch cell
# cell = sheet.cell(3,3)
# print('\nFetched Cell')
# pp.pprint(cell.value)


# #Update Cell
# cell = sheet.cell(3,3)
# print('Cell Before Update: ',cell.value)
# sheet.update_cell(3,3,'N')
# cell = sheet.cell(3,3)
# print('Cell After Update: ',cell.value)

# #Insert Row
# row = ['7','https://daily-py.blogspot.com','Y']
# index = 8
# sheet.insert_row(row,index)